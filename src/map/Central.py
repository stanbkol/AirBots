from datetime import datetime, timedelta
import json

import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from src.database.Models import *
from src.map.TileClassifier import *
from src.agents.Agents import *
from src.database.DataLoader import *
from sklearn.metrics import mean_squared_error, mean_absolute_error


def findNearestSensors(sensorid, s_list, n=0):
    """

    :param sensorid: sensor for which to find nearest neighbors.
    :param s_list: list of active sensors
    :return: list of (Sensor object, meters distance) tuples
    """
    base_sensor = getSensorORM(sensorid)
    sensors_orm = []

    for s in s_list:
        if s != sensorid:
            sensors_orm.append(getSensorORM(s))

    distances = []
    startLL = MapPoint(base_sensor.lat, base_sensor.lon)
    for sensor in sensors_orm:
        meters_away = calcDistance(startLL, MapPoint(sensor.lat, sensor.lon))
        distances.append((sensor, meters_away))

    distances.sort(key=lambda x: x[1])
    if n == 0:
        return distances
    else:
        return distances[:n]


def fetchBB(data):
    bounding_box = []
    for entry in data:
        bounding_box.append(entry["boundingbox"])
    return bounding_box[0]


# returns a dictionary containing the json file
def getJson(file):
    f = open(file, encoding="utf8")
    data = json.load(f)
    return data


# check database for all measurements data for given sensor, using different granularity (ie. for full dataset,
# by year, by month, by week, by day); should it be done with multiple, specific queries? or taking the full dataset
# for a sensor and breaking it down using additional algorithms
def examineSensor(sid, g="*"):
    full_dataset = getMeasuresORM(sid)
    earliest_date = full_dataset[0].date
    latest_date = full_dataset[-1].date
    total = countInterval(earliest_date, latest_date)
    observed = len(full_dataset)
    summary = {"sensor": sid, "first": earliest_date.strftime("%D %H"), "last": latest_date.strftime("%D %H"),
               "num_intervals": observed,
               "total": countInterval(datetime.datetime(2018, 9, 3, 0), datetime.datetime(2021, 5, 5, 0))}
    return summary


def countInterval(start, end):
    diff = end - start
    days, seconds = diff.days, diff.seconds
    total_intervals = days * 24 + seconds // 3600
    return total_intervals


def rateInterval(dataset, total):
    return (dataset - 1) / total


def fetchSensors(sensor_list):
    sensors = []
    for s in sensor_list:
        sensors.append(getSensorORM(s))
    return sensors


def classifyTiles(filename):
    data = getJson(filename)
    for tile in data:
        print("Classifying Tile:", tile)
        tile_class = classifyT(data[tile])
        print("Class-->", tile_class)
        tile_obj = getTileORM(tile)
        tile_obj.setClass(tile_class)


def p_err(a, p):
    num_preds = len(a)
    total = 0
    for i in range(0, num_preds):
        temp = abs((p[i] - a[i]) / a[i])
        total += temp
    return round((total / num_preds) * 100, 2)


def MSE(a, p):
    actual, pred = np.array(a), np.array(p)
    return round(mean_squared_error(actual, pred), 2)


def MAE(a, p):
    actual, pred = np.array(a), np.array(p)
    return round(mean_absolute_error(actual, pred), 2)


def avgAgg(preds):
    total = 0
    for k in preds:
        total += preds[k]
    return round(total / (len(preds)), 2)


# prediction weighted based on distance from each agent to the target
def WeightedAggregation(preds, weights):
    total = 0
    for s in preds:
        temp = preds[s]
        total += temp * weights.get(s)
    return total


def totalDist(data):
    total = 0
    for e in data:
        total += e[1]
    return total


def totalError(data):
    total = 0
    for a in data:
        agent = data[a]
        total += agent.error
    return total


# def distWeights(target, sensors):
#     weights = {}
#     dist_list = findNearestSensors(target, sensors)
#     total = totalDist(dist_list)
#     for s in dist_list:
#         weights[s[0].sid] = (s[1]/total)
#     print("Distance Weights", weights)
#     return weights

def kriegWeights(data, pow=2):
    """
    calculates Inverse Distance Weights for known sensors relative to an unmeasured target.
    :param target: unmeasured target
    :param sensors: known points for which weights are made
    :param pow: determines the rate at which the weights decrease. Default is 2.
    :return: dict of sid-->weight on scale between 0 and 1.
    """
    sumsup = [(x[0], 1 / np.power(x[1], pow)) for x in data]
    suminf = sum(n for _, n in sumsup)
    weights = {x[0].sid: x[1] / suminf for x in sumsup}
    return weights


def mapAgentsToError(agents):
    data = []
    for a in agents:
        data.append((agents[a], agents[a].error))
    data.sort(key=lambda x: x[1])
    return data


def checkWeights(w):
    total = 0
    for x in w:
        total += w[x]
    return total


# update to include other aggregation methods
# calc distance and error weights in main body and pass the values into agg methods where needed


def aggregatePrediction(preds, dist_weights, error_weights, tru_weights):
    vals = {}
    avg = avgAgg(preds)
    dist = WeightedAggregation(preds, dist_weights)
    err = WeightedAggregation(preds, error_weights)
    # trust = trustAgg(preds, tru_weights)
    vals['average'] = round(avg, 2)
    vals['distance'] = round(dist, 2)
    vals['error'] = round(err, 2)
    return vals


def makeCluster(sid, sensors, n):
    data = findNearestSensors(sid, sensors, n)
    cluster = []
    for sens, dist in data:
        cluster.append(sens.sid)
    return cluster


def getClusterError(cluster, agents):
    total = 0
    count = 0
    for a in cluster:
        total += agents[a].n_error
        count += 1
    return round(total / count, 2)


def targetInterval(time, interval):
    end = time + timedelta(hours=interval)
    start = time - timedelta(hours=interval)
    return start, end


class Central:
    agents = {}
    sensors = []
    agent_configs = {}
    thresholds = {}

    def __init__(self, model):
        self.model_file = model
        self.results_file = self.model_file + "_results.xlsx"
        self.data = getJson(self.model_file)
        self.model_params = self.data["model_params"]
        self.thresholds = self.data["thresholds"]
        self.error = 0
        self.popSensors()
        self.extractData()
        self.initializeFile()

    def evaluateAgents(self, values, predictions, naive_preds):
        # print("values:")
        # print(values)
        for a in self.agents:
            agent = self.agents[a]
            # print("predictions for:", a)
            # print(predictions[a])
            agent.error = MAE(values, predictions[a])
            agent.n_error = MAE(values, naive_preds[a])
            agent.p_error = (p_err(values, naive_preds[a]), p_err(values, predictions[a]))

    def popSensors(self):
        for s in self.data["sensors"]["on"]:
            self.sensors.append(s)

    def getAllPredictions(self, target, time, val):
        predictions = {}
        for a in self.agents:
            agent = self.agents[a]
            pred = agent.makePredictions(target, time, ["pm1"], meas=val)
            predictions[a] = (round(pred[0], 2), agent.cf)
        return predictions

    def makePrediction(self, target, time):
        start, end = targetInterval(time, self.thresholds["interval"])
        print("Training the Model")
        print("Target Sensor:", target)
        print("Interval between " + str(start) + " and " + str(end))
        self.trainModel(start, end, target)
        # predictions = []
        # for a in self.agents:
        #     agent = self.agents[a]
        #     p = agent.makePrediction(target, time)
        #     predictions.append(p)
        # return predictions

    def trainModel(self, start_interval, end_interval, target):
        for i in range(1, self.model_params["num_iter"] + 1):
            cursor = start_interval
            collab_predictions = {sid: [] for sid in self.sensors}
            naive_predictions = {sid: [] for sid in self.sensors}
            values = []
            intervals = []
            while cursor != end_interval:
                print("Predictions for ", cursor)
                val = getMeasureORM(target, cursor)
                # print("Hard Value-->", val.pm1)
                intervals.append(cursor)
                vals = {sid: [] for sid in self.sensors}
                if val:
                    values.append(val.pm1)
                    interval_preds = self.getAllPredictions(target, cursor, val)
                    # print(interval_preds)
                    for a in self.agents:
                        cluster_pred = {}
                        agent = self.agents[a]
                        for ca in agent.cluster:
                            cluster_pred[ca] = interval_preds[ca]
                        pred = round(agent.makeCollabPrediction(cluster_pred)[0], 2)
                        naive = interval_preds[a][0]
                        vals[a] = pred
                        collab_predictions[a].append(pred)
                        naive_predictions[a].append(naive)
                        # print("Agent: ", a)
                        # print("Naive Prediction: ", interval_preds[a][0])
                        # print("Prediction Cluster: ", cluster_pred)
                        # print("Collab Prediction: ", pred)
                else:
                    print("No target validation data for-->", cursor)
                cursor += timedelta(hours=1)

            self.evaluateAgents(values, collab_predictions, naive_predictions)
            model_vals = self.aggregateModel(collab_predictions, countInterval(start_interval, end_interval), target)
            self.evaluateModel(values, model_vals)
            self.saveIter(values, collab_predictions, naive_predictions, model_vals, i, intervals)
            self.saveModel(i)
            self.applyHeuristic(values, naive_predictions, collab_predictions, intervals)

    def sensorSummary(self):
        for s in self.sensors:
            print(examineSensor(s))

    def evaluateModel(self, values, model_preds):
        error = {}
        avg_list = []
        dist_list = []
        err_list = []
        for e in model_preds:
            avg_list.append(e['average'])
            dist_list.append(e['distance'])
            err_list.append(e['error'])
        error['MAE'] = {'average': MAE(values, avg_list), 'dist_w': MAE(values, dist_list),
                        'error_w': MAE(values, err_list)}
        error['p'] = {'average': p_err(values, avg_list), 'dist_w': p_err(values, dist_list),
                      'error_w': p_err(values, err_list)}
        print(error)
        self.error = error

    def extractData(self):
        self.thresholds = self.data["thresholds"]
        self.agent_configs = self.data["agent_configs"]
        for s in self.sensors:
            cluster = makeCluster(s, self.sensors, self.model_params['cluster_size'])
            a = Agent(s, self.thresholds, cluster, config=self.agent_configs)
            self.agents[a.sid] = a

    def aggregateModel(self, preds, num_preds, target):
        model_vals = []
        dist_weights = kriegWeights(findNearestSensors(target, self.sensors))
        err_weights = kriegWeights(mapAgentsToError(self.agents))
        tru_weights = {}
        for i in range(0, num_preds):
            interval_preds = {}
            for a in preds:
                interval_preds[a] = preds[a][i]
            model_vals.append(aggregatePrediction(interval_preds, dist_weights, err_weights, tru_weights))
        return model_vals

    def saveModel(self, i):
        print("Saving Data-->Iter #", i)
        wb = load_workbook(self.results_file)
        ws = wb["Training_Results"]
        col = i * 2
        row = 1
        ws.cell(row, col, "I" + str(i) + " (MAE)")
        ws.cell(row, col + 1, "I" + str(i) + " (%)")

        row_n = 2
        row_c = 3
        for a in self.agents:
            agent = self.agents[a]
            print("Error for :", a)
            print("Naive-->" + str(agent.n_error) + ":" + str(agent.p_error[0]))
            print("Collab-->" + str(agent.error) + ":" + str(agent.p_error[1]))
            ws.cell(row_n, col, agent.n_error)
            ws.cell(row_n, col + 1, agent.p_error[0])
            ws.cell(row_c, col, agent.error)
            ws.cell(row_c, col + 1, agent.p_error[1])
            row_n += 2
            row_c += 2
        ws.cell(row_n, col, self.error['MAE']['error_w'])
        ws.cell(row_n, col + 1, self.error['p']['error_w'])
        wb.save(self.results_file)

    def initializeFile(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Training_Results"
        ws.cell(1, 1, "Agents")
        row_n = 2
        row_c = 3
        for s in self.agents:
            ws.cell(row_n, 1, "N" + str(s))
            ws.cell(row_c, 1, s)
            row_c += 2
            row_n += 2
        ws.cell(row_n, 1, "Model")
        wb.save(filename=self.results_file)

    def saveIter(self, values, collab_predictions, naive_predictions, model_vals, i, intervals):
        wb = load_workbook(self.results_file)
        ws = wb.create_sheet("Iteration #" + str(i))
        col = 1
        row = 1
        ws.cell(row, col, "Values")
        ws.column_dimensions[get_column_letter(col)].width = 20
        # write time interval col headings
        for i in intervals:
            col += 1
            ws.cell(row, col, i)
            ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.cell(row, col, "Agent Bias")
        # write actual values, and 3x model aggregation values
        actual_ind = 2
        modelavg_ind = 3
        dist_ind = 4
        error_ind = 5
        num_i = len(intervals)
        print(num_i)
        ws.cell(actual_ind, 1, "Real Values")
        ws.cell(modelavg_ind, 1, "Model AVG")
        ws.cell(dist_ind, 1, "Model DIST")
        ws.cell(error_ind, 1, "Model ERR")
        for i in range(2, num_i + 2):
            ws.cell(actual_ind, i, values[i - 2])
            ws.cell(modelavg_ind, i, model_vals[i - 2]['average'])
            ws.cell(dist_ind, i, model_vals[i - 2]['distance'])
            ws.cell(error_ind, i, model_vals[i - 2]['error'])

        # write naive and collab values from each agent
        agent_n_index = 6
        agent_c_index = 7
        for a in collab_predictions:
            col = 1
            ws.cell(agent_n_index, col, "Agent #" + str(a) + ":Naive")
            ws.cell(agent_c_index, col, "Agent #" + str(a) + ":Collab")
            num = len(collab_predictions[a])
            for p in range(0, num):
                col += 1
                ws.cell(agent_n_index, col, naive_predictions[a][p])
                ws.cell(agent_c_index, col, collab_predictions[a][p])
            col += 1
            ws.cell(agent_c_index, col, self.agents[a].bias)
            agent_n_index += 2
            agent_c_index += 2
        wb.save(self.results_file)

    def applyHeuristic(self, values, naive_predictions, collab_predictions, intervals):
        for a in self.agents:
            print("agent H:", a)
            agent = self.agents[a]
            key_list = [a]
            key_list.extend(agent.cluster)
            n_preds = {}
            for k in key_list:
                n_preds[k] = naive_predictions[k]
            agent.improveHeuristic(values, n_preds, collab_predictions[a], intervals)
