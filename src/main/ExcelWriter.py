from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


class ExcelWriter:
    def __init__(self, file):
        self.results_file = file

    def saveModel(self, i, agents, model_error, title):
        wb = load_workbook(self.results_file)
        ws = wb["Training_Results"]
        col = i * 2
        row = 1
        ws.cell(row, col, title + str(i) + " (MAE)")
        ws.cell(row, col + 1, title + str(i) + " (%)")
        row_n = 2
        row_c = 3
        for a in agents:
            agent = agents[a]
            # print("Error for :", a)
            # print("Naive-->" + str(agent.n_error) + ":" + str(agent.p_error[0]))
            # print("Collab-->" + str(agent.error) + ":" + str(agent.p_error[1]))
            ws.cell(row_n, col, agent.n_error)
            ws.cell(row_n, col + 1, agent.p_error[0])
            ws.cell(row_c, col, agent.error)
            ws.cell(row_c, col + 1, agent.p_error[1])
            row_n += 2
            row_c += 2
        ws.cell(row_n, col, model_error['MAE']['average'])
        ws.cell(row_n, col + 1, model_error['p']['average'])
        ws.cell(row_n + 1, col, model_error['MAE']['error_w'])
        ws.cell(row_n + 1, col + 1, model_error['p']['error_w'])
        ws.cell(row_n + 2, col, model_error['MAE']['trust_w'])
        ws.cell(row_n + 2, col + 1, model_error['p']['trust_w'])
        wb.save(self.results_file)

    def initializeFile(self, agents):
        wb = Workbook()
        ws = wb.active
        ws.title = "Training_Results"
        ws.cell(1, 1, "Agents")
        row_n = 2
        row_c = 3
        for s in agents:
            ws.cell(row_n, 1, "N" + str(s))
            ws.cell(row_c, 1, s)
            row_c += 2
            row_n += 2
        ws.cell(row_n, 1, "Model AVG")
        ws.cell(row_n + 1, 1, "Model ERR")
        ws.cell(row_n + 2, 1, "Model TF")
        wb.save(filename=self.results_file)

    def saveIter(self, values, collab_predictions, naive_predictions, model_vals, i, intervals, agents):
        wb = load_workbook(self.results_file)
        ws = wb.create_sheet("Iteration #"+str(i))
        col = 1
        row = 1
        ws.cell(row, col, "Values")
        ws.column_dimensions[get_column_letter(col)].width = 20
        # write time interval col headings
        for i in intervals:
            col += 1
            ws.cell(row, col, i)
            ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        bias_col = col
        nearby_col = col + 1
        minmax_col = col + 2
        sma_col = col + 3
        mvr_col = col + 4
        ws.cell(row, bias_col, "Agent Bias")
        ws.cell(row, nearby_col, "Nearby-->N")
        ws.cell(row, minmax_col, "MinMax-->N")
        ws.cell(row, sma_col, "SMA-->window")
        ws.cell(row, mvr_col, "MVR-->interval_hours")
        ws.column_dimensions[get_column_letter(bias_col)].width = 20
        ws.column_dimensions[get_column_letter(nearby_col)].width = 20
        ws.column_dimensions[get_column_letter(minmax_col)].width = 20
        ws.column_dimensions[get_column_letter(sma_col)].width = 20
        ws.column_dimensions[get_column_letter(mvr_col)].width = 20

        actual_ind = 2
        modelavg_ind = 3
        error_ind = 4
        tf_ind = 5
        num_i = len(intervals)
        ws.cell(actual_ind, 1, "Real Values")
        ws.cell(modelavg_ind, 1, "Model AVG")
        ws.cell(error_ind, 1, "Model ERR")
        ws.cell(tf_ind, 1, "Model TF")
        for i in range(2, num_i + 2):
            ws.cell(actual_ind, i, values[i - 2])
            ws.cell(modelavg_ind, i, model_vals[i - 2]['average'])
            ws.cell(error_ind, i, model_vals[i - 2]['error'])
            ws.cell(tf_ind, i, model_vals[i - 2]['trust'])

        # write naive and collab values from each agent
        agent_n_index = 6
        agent_c_index = 7
        for a in collab_predictions:
            col = 1
            ws.cell(agent_n_index, col, "Agent #" + str(a) + ":Naive")
            ws.cell(agent_c_index, col, "Agent #" + str(a) + ":Collab")
            num = len(collab_predictions[a])
            for p in range(0, num):
                col += 1
                ws.cell(agent_n_index, col, naive_predictions[a][p])
                ws.cell(agent_c_index, col, collab_predictions[a][p])
            col += 1
            bias_col = col
            nearby_col = col + 1
            minmax_col = col + 2
            sma_col = col + 3
            mvr_col = col + 4
            ws.cell(agent_n_index, bias_col, agents[a].configs['bias'])
            ws.cell(agent_n_index, nearby_col, agents[a].configs['nearby']['n'])
            ws.cell(agent_n_index, minmax_col, agents[a].configs['minmax']['n'])
            ws.cell(agent_n_index, sma_col, str(agents[a].configs['sma']['window']) + " : " + str(agents[a].configs['sma']['interval_hours']))
            ws.cell(agent_n_index, mvr_col, agents[a].configs['mvr']['interval_hours'])
            agent_n_index += 2
            agent_c_index += 2
        wb.save(self.results_file)

    def saveAgentConfigs(self, agent_results):
        wb = load_workbook(self.results_file)
        ws = wb.create_sheet("Selfish_Agent_Summary")
        ws.cell(1, 1, "Selfish Configs")
        bias_i = 2
        nearby_i = 3
        minmax_i = 4
        sma_i = 5
        mvr_i = 6
        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.cell(bias_i, 1, "Bias")
        ws.cell(nearby_i, 1, "Nearby-->n")
        ws.cell(minmax_i, 1, "MinMax-->n")
        ws.cell(sma_i, 1, "SMA-->window")
        ws.cell(mvr_i, 1, "MVR-->interval_hours")
        agent_c = 2
        for a in agent_results:
            ws.cell(1, agent_c, a)
            ws.cell(bias_i, agent_c, agent_results[a]['config']['bias'])
            ws.cell(nearby_i, agent_c, agent_results[a]['config']['nearby']['n'])
            ws.cell(minmax_i, agent_c, agent_results[a]['config']['minmax']['n'])
            ws.cell(sma_i, agent_c, str(agent_results[a]['config']['sma']['window']) + " : " + str(agent_results[a]['config']['sma']['interval_hours']))
            ws.cell(mvr_i, agent_c, agent_results[a]['config']['mvr']['interval_hours'])
            agent_c += 1
        wb.save(self.results_file)

    def saveModelBest(self, model_results):
        wb = load_workbook(self.results_file)
        ws = wb.create_sheet("Historical_Agent_Summary")
        ws.cell(1, 1, "Historical Configs")
        ws.column_dimensions[get_column_letter(1)].width = 20
        bias_i = 2
        nearby_i = 3
        minmax_i = 4
        sma_i = 5
        mvr_i = 6
        ws.cell(bias_i, 1, "Bias")
        ws.cell(nearby_i, 1, "Nearby-->n")
        ws.cell(minmax_i, 1, "MinMax-->n")
        ws.cell(sma_i, 1, "SMA-->window")
        ws.cell(mvr_i, 1, "MVR-->interval_hours")
        agent_c = 2
        for a in model_results:
            ws.cell(1, agent_c, a)
            ws.cell(bias_i, agent_c, model_results[a]['bias'])
            ws.cell(nearby_i, agent_c, model_results[a]['nearby']['n'])
            ws.cell(minmax_i, agent_c, model_results[a]['minmax']['n'])
            ws.cell(sma_i, agent_c, str(model_results[a]['sma']['window']) + " : " + str(model_results[a]['sma']['interval_hours']))
            ws.cell(mvr_i, agent_c, model_results[a]['mvr']['interval_hours'])
            agent_c += 1
        wb.save(self.results_file)